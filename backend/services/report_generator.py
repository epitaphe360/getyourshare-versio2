"""
Générateur de rapports PDF, CSV et Excel
Service complet pour exports de données
"""

import os
import csv
import json
from io import BytesIO, StringIO
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
from enum import Enum
from utils.logger import logger

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.info("⚠️ reportlab pas installé - Génération PDF désactivée")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference, LineChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.info("⚠️ openpyxl pas installé - Génération Excel désactivée")


class ReportType(str, Enum):
    """Types de rapports"""
    REVENUE = "revenue"
    CONVERSIONS = "conversions"
    AFFILIATES = "affiliates"
    COMMISSIONS = "commissions"
    PRODUCTS = "products"
    ANALYTICS = "analytics"


class ReportFormat(str, Enum):
    """Formats d'export"""
    PDF = "pdf"
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"


class ReportGenerator:
    """Service de génération de rapports"""
    
    def __init__(self):
        self.reports_dir = os.path.join(os.path.dirname(__file__), "..", "reports")
        os.makedirs(self.reports_dir, exist_ok=True)
        
        self.pdf_available = REPORTLAB_AVAILABLE
        self.excel_available = OPENPYXL_AVAILABLE
    
    def generate_report(
        self,
        report_type: ReportType,
        format: ReportFormat,
        data: Dict[str, Any],
        user_info: Dict[str, Any],
        filters: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Générer un rapport dans le format demandé
        
        Args:
            report_type: Type de rapport
            format: Format d'export (PDF, CSV, Excel, JSON)
            data: Données du rapport
            user_info: Informations utilisateur
            filters: Filtres appliqués
        
        Returns:
            Chemin du fichier généré et métadonnées
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type.value}_{timestamp}"
        
        if format == ReportFormat.PDF:
            return self._generate_pdf(filename, report_type, data, user_info, filters)
        elif format == ReportFormat.CSV:
            return self._generate_csv(filename, report_type, data)
        elif format == ReportFormat.EXCEL:
            return self._generate_excel(filename, report_type, data, user_info, filters)
        elif format == ReportFormat.JSON:
            return self._generate_json(filename, report_type, data, filters)
        else:
            raise ValueError(f"Format non supporté: {format}")
    
    def _generate_pdf(
        self,
        filename: str,
        report_type: ReportType,
        data: Dict[str, Any],
        user_info: Dict[str, Any],
        filters: Optional[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Générer un rapport PDF professionnel"""
        if not self.pdf_available:
            return {
                "success": False,
                "error": "reportlab n'est pas installé. Utilisez: pip install reportlab"
            }
        
        filepath = os.path.join(self.reports_dir, f"{filename}.pdf")
        
        # Créer le document
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Style personnalisé pour le titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # En-tête
        title = f"Rapport {report_type.value.upper()}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Informations utilisateur et période
        info_data = [
            ['Entreprise:', user_info.get('company_name', 'N/A')],
            ['Généré par:', user_info.get('name', 'N/A')],
            ['Date:', datetime.now().strftime('%d/%m/%Y %H:%M')],
            ['Période:', filters.get('period', 'Tous les temps') if filters else 'Tous les temps']
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Métriques principales selon le type de rapport
        if report_type == ReportType.REVENUE:
            metrics = self._add_revenue_metrics(data, elements, styles)
        elif report_type == ReportType.CONVERSIONS:
            metrics = self._add_conversion_metrics(data, elements, styles)
        elif report_type == ReportType.COMMISSIONS:
            metrics = self._add_commission_metrics(data, elements, styles)
        elif report_type == ReportType.AFFILIATES:
            metrics = self._add_affiliate_metrics(data, elements, styles)
        else:
            metrics = self._add_generic_metrics(data, elements, styles)
        
        # Construire le PDF
        doc.build(elements)
        
        return {
            "success": True,
            "filepath": filepath,
            "filename": f"{filename}.pdf",
            "format": "PDF",
            "size_bytes": os.path.getsize(filepath),
            "generated_at": datetime.now().isoformat()
        }
    
    def _add_revenue_metrics(self, data: Dict, elements: List, styles) -> None:
        """Ajouter les métriques de revenus au PDF"""
        # Titre section
        elements.append(Paragraph("📊 Résumé des Revenus", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Tableau des métriques clés
        metrics_data = [
            ['Métrique', 'Valeur'],
            ['Revenu Total', f"{data.get('total_revenue', 0):,.2f} MAD"],
            ['Nombre de Ventes', str(data.get('total_sales', 0))],
            ['Panier Moyen', f"{data.get('average_order', 0):,.2f} MAD"],
            ['Taux de Conversion', f"{data.get('conversion_rate', 0):.2f}%"],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 3*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Top produits
        if 'top_products' in data:
            elements.append(Paragraph("🏆 Top 5 Produits", styles['Heading3']))
            elements.append(Spacer(1, 0.1*inch))
            
            products_data = [['Produit', 'Ventes', 'Revenu']]
            for product in data['top_products'][:5]:
                products_data.append([
                    product['name'],
                    str(product['sales']),
                    f"{product['revenue']:,.2f} MAD"
                ])
            
            products_table = Table(products_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            products_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(products_table)
    
    def _add_conversion_metrics(self, data: Dict, elements: List, styles) -> None:
        """Ajouter les métriques de conversions"""
        elements.append(Paragraph("🎯 Entonnoir de Conversion", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        funnel_data = [
            ['Étape', 'Utilisateurs', 'Taux'],
            ['Visites', str(data.get('visits', 0)), '100%'],
            ['Clics', str(data.get('clicks', 0)), f"{data.get('click_rate', 0):.1f}%"],
            ['Paniers', str(data.get('carts', 0)), f"{data.get('cart_rate', 0):.1f}%"],
            ['Achats', str(data.get('purchases', 0)), f"{data.get('purchase_rate', 0):.1f}%"],
        ]
        
        funnel_table = Table(funnel_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        funnel_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 11),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightblue, colors.white]),
        ]))
        elements.append(funnel_table)
    
    def _add_commission_metrics(self, data: Dict, elements: List, styles) -> None:
        """Ajouter les métriques de commissions"""
        elements.append(Paragraph("💰 Commissions Payées", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Liste des commissions
        if 'commissions' in data:
            comm_data = [['Influenceur', 'Ventes', 'Commission', 'Statut']]
            for comm in data['commissions']:
                comm_data.append([
                    comm['influencer_name'],
                    str(comm['sales_count']),
                    f"{comm['amount']:,.2f} MAD",
                    comm['status']
                ])
            
            comm_table = Table(comm_data, colWidths=[2.5*inch, 1.2*inch, 1.5*inch, 1.3*inch])
            comm_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(comm_table)
    
    def _add_affiliate_metrics(self, data: Dict, elements: List, styles) -> None:
        """Ajouter les métriques des affiliés"""
        elements.append(Paragraph("👥 Performance des Affiliés", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        if 'affiliates' in data:
            aff_data = [['Affilié', 'Clics', 'Ventes', 'Taux Conv.', 'Revenu']]
            for aff in data['affiliates']:
                aff_data.append([
                    aff['name'],
                    str(aff['clicks']),
                    str(aff['sales']),
                    f"{aff['conversion_rate']:.1f}%",
                    f"{aff['revenue']:,.2f} MAD"
                ])
            
            aff_table = Table(aff_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1.5*inch])
            aff_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ec4899')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(aff_table)
    
    def _add_generic_metrics(self, data: Dict, elements: List, styles) -> None:
        """Ajouter des métriques génériques"""
        elements.append(Paragraph("📈 Données du Rapport", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Convertir le dictionnaire en tableau
        generic_data = [['Métrique', 'Valeur']]
        for key, value in data.items():
            if not isinstance(value, (dict, list)):
                generic_data.append([str(key).replace('_', ' ').title(), str(value)])
        
        generic_table = Table(generic_data, colWidths=[3*inch, 3*inch])
        generic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(generic_table)
    
    def _generate_csv(
        self,
        filename: str,
        report_type: ReportType,
        data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Générer un rapport CSV"""
        filepath = os.path.join(self.reports_dir, f"{filename}.csv")
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            # Déterminer les colonnes selon le type de rapport
            if report_type == ReportType.REVENUE:
                fieldnames = ['date', 'revenue', 'orders', 'average_order']
                rows = data.get('daily_revenue', [])
            elif report_type == ReportType.CONVERSIONS:
                fieldnames = ['date', 'visits', 'clicks', 'conversions', 'rate']
                rows = data.get('daily_conversions', [])
            elif report_type == ReportType.COMMISSIONS:
                fieldnames = ['influencer', 'sales', 'commission', 'status']
                rows = data.get('commissions', [])
            else:
                # Export générique
                if data and isinstance(data, list):
                    if len(data) > 0 and isinstance(data[0], dict):
                        fieldnames = list(data[0].keys())
                        rows = data
                    else:
                        fieldnames = ['value']
                        rows = [{'value': item} for item in data]
                else:
                    fieldnames = ['key', 'value']
                    rows = [{'key': k, 'value': v} for k, v in data.items() if not isinstance(v, (dict, list))]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        
        return {
            "success": True,
            "filepath": filepath,
            "filename": f"{filename}.csv",
            "format": "CSV",
            "rows": len(rows) if rows else 0,
            "size_bytes": os.path.getsize(filepath),
            "generated_at": datetime.now().isoformat()
        }
    
    def _generate_excel(
        self,
        filename: str,
        report_type: ReportType,
        data: Dict[str, Any],
        user_info: Dict[str, Any],
        filters: Optional[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Générer un rapport Excel avec graphiques"""
        if not self.excel_available:
            return {
                "success": False,
                "error": "openpyxl n'est pas installé. Utilisez: pip install openpyxl"
            }
        
        filepath = os.path.join(self.reports_dir, f"{filename}.xlsx")
        
        # Créer le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.value.title()
        
        # Styles
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # En-tête du rapport
        ws['A1'] = f"Rapport {report_type.value.upper()}"
        ws['A1'].font = Font(size=16, bold=True, color="2563eb")
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'] = f"Entreprise: {user_info.get('company_name', 'N/A')}"
        
        # Ajouter les données selon le type
        start_row = 5
        if report_type == ReportType.REVENUE:
            headers = ['Date', 'Revenu', 'Commandes', 'Panier Moyen']
            ws.append([''] * len(headers))  # Ligne vide
            ws.append(headers)
            
            # Appliquer le style aux en-têtes
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=start_row + 1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            # Ajouter les données
            for row_data in data.get('daily_revenue', []):
                ws.append([
                    row_data.get('date'),
                    row_data.get('revenue', 0),
                    row_data.get('orders', 0),
                    row_data.get('average_order', 0)
                ])
        
        # Ajuster les largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder
        wb.save(filepath)
        
        return {
            "success": True,
            "filepath": filepath,
            "filename": f"{filename}.xlsx",
            "format": "Excel",
            "size_bytes": os.path.getsize(filepath),
            "generated_at": datetime.now().isoformat()
        }
    
    def _generate_json(
        self,
        filename: str,
        report_type: ReportType,
        data: Dict[str, Any],
        filters: Optional[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Générer un rapport JSON"""
        filepath = os.path.join(self.reports_dir, f"{filename}.json")
        
        report_data = {
            "report_type": report_type.value,
            "generated_at": datetime.now().isoformat(),
            "filters": filters or {},
            "data": data
        }
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(report_data, jsonfile, indent=2, ensure_ascii=False)
        
        return {
            "success": True,
            "filepath": filepath,
            "filename": f"{filename}.json",
            "format": "JSON",
            "size_bytes": os.path.getsize(filepath),
            "generated_at": datetime.now().isoformat()
        }


# Instance singleton
report_generator = ReportGenerator()
