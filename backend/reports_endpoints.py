"""
============================================
ADVANCED REPORTS ENDPOINTS
GetYourShare - Système de Rapports Avancés
============================================

Endpoints pour génération et export de rapports:
- Rapports ventes, commissions, clics
- Export multi-format (CSV, Excel, PDF)
- Programmation de rapports
- Comparaisons temporelles
"""

from fastapi import APIRouter, HTTPException, Depends, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from auth import get_current_user
from supabase_client import supabase
from utils.logger import logger
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/api/reports", tags=["Reports"])

# ============================================
# PYDANTIC MODELS
# ============================================

class ScheduleReportRequest(BaseModel):
    """Programmation de rapport"""
    report_type: str
    frequency: str  # daily, weekly, monthly
    format: str  # csv, xlsx, pdf
    recipients: List[str] = []

# ============================================
# HELPER FUNCTIONS
# ============================================

def calculate_period_comparison(start_date: datetime, end_date: datetime, period_type: str):
    """Calcule les dates de comparaison"""
    delta = end_date - start_date
    
    if period_type == "previous":
        comparison_start = start_date - delta
        comparison_end = start_date
    elif period_type == "last_year":
        comparison_start = start_date - timedelta(days=365)
        comparison_end = end_date - timedelta(days=365)
    else:
        return None, None
    
    return comparison_start, comparison_end


def get_sales_report_data(user_id: str, user_role: str, start_date: str, end_date: str):
    """Génère les données du rapport de ventes"""
    try:
        # Base query
        query = supabase.table('transactions').select('''
            id,
            amount,
            commission,
            status,
            created_at,
            products(name),
            users(first_name, last_name)
        ''').gte('created_at', start_date).lte('created_at', end_date)
        
        # Filtrer selon le rôle
        if user_role == 'merchant':
            query = query.eq('merchant_id', user_id)
        elif user_role == 'influencer':
            query = query.eq('influencer_id', user_id)
        
        response = query.execute()
        transactions = response.data if response.data else []
        
        # Calculer les stats
        total_revenue = sum([t.get('amount', 0) for t in transactions])
        total_commission = sum([t.get('commission', 0) for t in transactions])
        total_orders = len(transactions)
        
        # Grouper par date pour le graphique
        chart_data = {}
        for t in transactions:
            date = t['created_at'][:10]
            if date not in chart_data:
                chart_data[date] = {'date': date, 'revenue': 0, 'orders': 0}
            chart_data[date]['revenue'] += t.get('amount', 0)
            chart_data[date]['orders'] += 1
        
        return {
            'transactions': transactions,
            'chart_data': list(chart_data.values()),
            'stats': {
                'total_revenue': total_revenue,
                'total_orders': total_orders,
                'total_commission': total_commission,
                'avg_order_value': total_revenue / total_orders if total_orders > 0 else 0
            }
        }
    except Exception as e:
        logger.error(f"Error generating sales report: {e}")
        return {'transactions': [], 'chart_data': [], 'stats': {}}


def get_commissions_report_data(user_id: str, user_role: str, start_date: str, end_date: str):
    """Génère les données du rapport de commissions"""
    try:
        query = supabase.table('commissions').select('''
            id,
            amount,
            status,
            created_at,
            users(first_name, last_name)
        ''').gte('created_at', start_date).lte('created_at', end_date)
        
        if user_role == 'influencer':
            query = query.eq('influencer_id', user_id)
        elif user_role == 'merchant':
            query = query.eq('merchant_id', user_id)
        
        response = query.execute()
        commissions = response.data if response.data else []
        
        # Stats
        earned = sum([c.get('amount', 0) for c in commissions])
        paid = sum([c.get('amount', 0) for c in commissions if c.get('status') == 'paid'])
        pending = sum([c.get('amount', 0) for c in commissions if c.get('status') == 'pending'])
        
        # Chart data
        chart_data = {}
        for c in commissions:
            date = c['created_at'][:10]
            if date not in chart_data:
                chart_data[date] = {'date': date, 'earned': 0, 'paid': 0, 'pending': 0}
            chart_data[date]['earned'] += c.get('amount', 0)
            if c.get('status') == 'paid':
                chart_data[date]['paid'] += c.get('amount', 0)
            elif c.get('status') == 'pending':
                chart_data[date]['pending'] += c.get('amount', 0)
        
        return {
            'commissions': commissions,
            'chart_data': list(chart_data.values()),
            'stats': {
                'total_earned': earned,
                'total_paid': paid,
                'total_pending': pending
            }
        }
    except Exception as e:
        logger.error(f"Error generating commissions report: {e}")
        return {'commissions': [], 'chart_data': [], 'stats': {}}


def get_clicks_report_data(user_id: str, user_role: str, start_date: str, end_date: str):
    """Génère les données du rapport de clics"""
    try:
        query = supabase.table('clicks').select('*').gte('created_at', start_date).lte('created_at', end_date)
        
        if user_role == 'influencer':
            query = query.eq('influencer_id', user_id)
        elif user_role == 'merchant':
            # Get merchant's campaigns
            campaigns_result = supabase.table('campaigns').select('id').eq('merchant_id', user_id).execute()
            campaign_ids = [c['id'] for c in (campaigns_result.data or [])]
            if campaign_ids:
                query = query.in_('campaign_id', campaign_ids)
        
        response = query.execute()
        clicks = response.data if response.data else []
        
        # Conversions
        conversions = [c for c in clicks if c.get('converted')]
        conversion_rate = (len(conversions) / len(clicks) * 100) if clicks else 0
        
        # Chart data
        chart_data = {}
        for c in clicks:
            date = c['created_at'][:10]
            if date not in chart_data:
                chart_data[date] = {'date': date, 'clicks': 0, 'conversions': 0}
            chart_data[date]['clicks'] += 1
            if c.get('converted'):
                chart_data[date]['conversions'] += 1
        
        return {
            'clicks': clicks,
            'chart_data': list(chart_data.values()),
            'stats': {
                'total_clicks': len(clicks),
                'total_conversions': len(conversions),
                'conversion_rate': conversion_rate
            }
        }
    except Exception as e:
        logger.error(f"Error generating clicks report: {e}")
        return {'clicks': [], 'chart_data': [], 'stats': {}}

# ============================================
# ENDPOINTS
# ============================================

@router.get("/generate")
async def generate_report(
    report_type: str = Query(..., regex="^(sales|commissions|clicks|affiliates|products|revenue)$"),
    start_date: str = Query(...),
    end_date: str = Query(...),
    comparison_period: str = Query("none"),
    current_user: Dict = Depends(get_current_user)
):
    """
    Génère un rapport selon le type demandé
    """
    try:
        user_id = current_user['id']
        user_role = current_user.get('role', 'user')
        
        # Générer le rapport selon le type
        if report_type == 'sales':
            report_data = get_sales_report_data(user_id, user_role, start_date, end_date)
        elif report_type == 'commissions':
            report_data = get_commissions_report_data(user_id, user_role, start_date, end_date)
        elif report_type == 'clicks':
            report_data = get_clicks_report_data(user_id, user_role, start_date, end_date)
        else:
            # Autres types de rapports (à implémenter)
            report_data = {'transactions': [], 'chart_data': [], 'stats': {}}
        
        # Comparaison avec période précédente si demandé
        if comparison_period != 'none':
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            comp_start, comp_end = calculate_period_comparison(start_dt, end_dt, comparison_period)
            
            if comp_start and comp_end:
                if report_type == 'sales':
                    comp_data = get_sales_report_data(
                        user_id, user_role,
                        comp_start.strftime('%Y-%m-%d'),
                        comp_end.strftime('%Y-%m-%d')
                    )
                    # Calculer les % de croissance
                    if comp_data['stats'].get('total_revenue'):
                        report_data['stats']['revenue_growth'] = round(
                            (report_data['stats']['total_revenue'] - comp_data['stats']['total_revenue']) / 
                            comp_data['stats']['total_revenue'] * 100, 2
                        )
                    if comp_data['stats'].get('total_orders'):
                        report_data['stats']['orders_growth'] = round(
                            (report_data['stats']['total_orders'] - comp_data['stats']['total_orders']) / 
                            comp_data['stats']['total_orders'] * 100, 2
                        )
        
        return {
            'success': True,
            'report': report_data,
            'chart_data': report_data.get('chart_data', []),
            'stats': report_data.get('stats', {}),
            'metadata': {
                'report_type': report_type,
                'start_date': start_date,
                'end_date': end_date,
                'generated_at': datetime.utcnow().isoformat()
            }
        }
    
    except Exception as e:
        logger.error(f"Error generating report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate report: {str(e)}"
        )


@router.get("/export")
async def export_report(
    report_type: str = Query(...),
    start_date: str = Query(...),
    end_date: str = Query(...),
    format: str = Query("csv", regex="^(csv|xlsx|pdf)$"),
    current_user: Dict = Depends(get_current_user)
):
    """
    Exporte un rapport au format demandé
    """
    try:
        user_id = current_user['id']
        user_role = current_user.get('role', 'user')
        
        # Générer les données
        if report_type == 'sales':
            report_data = get_sales_report_data(user_id, user_role, start_date, end_date)
            data_key = 'transactions'
        elif report_type == 'commissions':
            report_data = get_commissions_report_data(user_id, user_role, start_date, end_date)
            data_key = 'commissions'
        elif report_type == 'clicks':
            report_data = get_clicks_report_data(user_id, user_role, start_date, end_date)
            data_key = 'clicks'
        else:
            raise HTTPException(status_code=400, detail="Invalid report type")
        
        data = report_data.get(data_key, [])
        
        # Export CSV
        if format == 'csv':
            output = io.StringIO()
            if data:
                writer = csv.DictWriter(output, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
            
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=report_{report_type}.csv"}
            )
        
        # Export Excel
        elif format == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = f"Report {report_type}"
            
            # Style
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            if data:
                # Headers
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Data
                for row, item in enumerate(data, 2):
                    for col, key in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=item.get(key))
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=report_{report_type}.xlsx"}
            )
        
        # Export PDF
        elif format == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                import io as _io

                buffer = _io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                styles = getSampleStyleSheet()
                elements = []

                # Titre
                elements.append(Paragraph(f"Rapport : {report_type.replace('_', ' ').title()}", styles['Title']))
                elements.append(Spacer(1, 12))

                if data:
                    headers_list = list(data[0].keys())
                    table_data = [headers_list]
                    for row in data:
                        table_data.append([str(row.get(h, '')) for h in headers_list])

                    t = Table(table_data)
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ]))
                    elements.append(t)
                else:
                    elements.append(Paragraph("Aucune donnée disponible.", styles['Normal']))

                doc.build(elements)
                buffer.seek(0)
                return StreamingResponse(
                    buffer,
                    media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=report_{report_type}.pdf"}
                )
            except ImportError:
                raise HTTPException(status_code=501, detail="reportlab non installé. Utilisez le format CSV ou XLSX.")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export report: {str(e)}"
        )


@router.post("/schedule", status_code=status.HTTP_201_CREATED)
async def schedule_report(
    request: ScheduleReportRequest,
    current_user: Dict = Depends(get_current_user)
):
    """
    Programme l'envoi automatique d'un rapport
    """
    try:
        schedule_data = {
            'user_id': current_user['id'],
            'report_type': request.report_type,
            'frequency': request.frequency,
            'format': request.format,
            'recipients': request.recipients,
            'active': True,
            'created_at': datetime.utcnow().isoformat()
        }
        
        response = supabase.table('scheduled_reports').insert(schedule_data).execute()
        
        if not response.data:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to schedule report"
            )
        
        return {
            'success': True,
            'message': 'Report scheduled successfully',
            'schedule': response.data[0]
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error scheduling report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to schedule report: {str(e)}"
        )


@router.get("/scheduled")
async def get_scheduled_reports(
    current_user: Dict = Depends(get_current_user)
):
    """
    Récupère les rapports programmés de l'utilisateur
    """
    try:
        response = supabase.table('scheduled_reports')\
            .select('*')\
            .eq('user_id', current_user['id'])\
            .order('created_at', desc=True)\
            .execute()
        
        return {
            'schedules': response.data if response.data else []
        }
    
    except Exception as e:
        logger.error(f"Error fetching scheduled reports: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch scheduled reports: {str(e)}"
        )


@router.delete("/scheduled/{schedule_id}")
async def delete_scheduled_report(
    schedule_id: str,
    current_user: Dict = Depends(get_current_user)
):
    """
    Supprime un rapport programmé
    """
    try:
        response = supabase.table('scheduled_reports')\
            .delete()\
            .eq('id', schedule_id)\
            .eq('user_id', current_user['id'])\
            .execute()
        
        return {
            'success': True,
            'message': 'Scheduled report deleted successfully'
        }
    
    except Exception as e:
        logger.error(f"Error deleting scheduled report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete scheduled report: {str(e)}"
        )
